import re
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from langchain_core.documents import Document

from app.config import EMOJI_SHEET_NAME, EMOTIONS_WORKBOOK
from app.schemas import EmotionGroup, EmotionOption


class KnowledgeBaseError(RuntimeError):
    pass


@dataclass(frozen=True)
class EmotionRecord:
    group: str
    emotion: str
    reason: str
    advice: str
    services: str


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("Â ", " ")).strip()


def emoji_code_to_character(code: object) -> str:
    code_text = clean_cell(code)
    codepoints = re.findall(r"U\+([0-9A-Fa-f]{4,6})", code_text)
    if not codepoints:
        return ""

    try:
        emoji = "".join(chr(int(codepoint, 16)) for codepoint in codepoints)
    except (OverflowError, ValueError):
        return ""

    if not emoji or all(unicodedata.category(character).startswith("C") for character in emoji):
        return ""
    return emoji


def is_emotion_category_row(emotion: str, reason: str, advice: str, services: str) -> bool:
    return bool(re.fullmatch(r"\d+\.\s*[A-Z ]+", emotion)) and not (
        reason or advice or services
    )


def load_emotion_emojis(workbook_path: Path = EMOTIONS_WORKBOOK) -> dict[str, str]:
    if not workbook_path.is_file():
        return {}

    try:
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
    except Exception:
        return {}

    try:
        if EMOJI_SHEET_NAME not in workbook.sheetnames:
            return {}
        worksheet = workbook[EMOJI_SHEET_NAME]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    main_row_index = next(
        (
            index
            for index, row in enumerate(rows)
            if row and clean_cell(row[0]).casefold() == "main emotion"
        ),
        None,
    )
    if main_row_index is None:
        return {}

    emojis: dict[str, str] = {}
    main_row = rows[main_row_index]
    for column_index in range(1, len(main_row), 2):
        group = clean_cell(main_row[column_index])
        if not group:
            continue

        code_column_index = column_index + 1
        emoji = ""
        for row in rows[main_row_index:]:
            if code_column_index >= len(row):
                continue
            emoji = emoji_code_to_character(row[code_column_index])
            if emoji:
                break

        if emoji:
            emojis[group.casefold()] = emoji

    return emojis


def load_emotion_records(workbook_path: Path = EMOTIONS_WORKBOOK) -> list[EmotionRecord]:
    if not workbook_path.is_file():
        raise KnowledgeBaseError(
            f"Required emotions workbook is missing: {workbook_path.name}"
        )

    try:
        workbook = openpyxl.load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise KnowledgeBaseError(
            f"Could not read emotions workbook: {workbook_path.name}"
        ) from exc

    records: list[EmotionRecord] = []
    try:
        for worksheet in workbook.worksheets:
            group = clean_cell(worksheet.title)
            if group.casefold() == EMOJI_SHEET_NAME.casefold():
                continue
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                emotion, reason, advice, services = (
                    clean_cell(row[index]) if index < len(row) else ""
                    for index in range(4)
                )
                if not emotion or emotion.casefold() == "emotion":
                    continue
                if is_emotion_category_row(emotion, reason, advice, services):
                    continue
                if not (reason or advice or services):
                    continue
                records.append(
                    EmotionRecord(
                        group=group,
                        emotion=emotion,
                        reason=reason,
                        advice=advice,
                        services=services,
                    )
                )
    finally:
        workbook.close()

    if not records:
        raise KnowledgeBaseError("The emotions workbook contains no usable rows.")
    return records


def emotion_records_to_documents(
    records: list[EmotionRecord],
    source_name: str = EMOTIONS_WORKBOOK.name,
) -> list[Document]:
    return [
        Document(
            page_content=(
                f"Emotion group: {record.group}\n"
                f"Emotion: {record.emotion}\n"
                f"Reason: {record.reason}\n"
                f"Advice: {record.advice}\n"
                f"Services: {record.services}"
            ),
            metadata={
                "source": source_name,
                "sheet": record.group,
                "emotion": record.emotion,
                "reason": record.reason,
                "services": record.services,
            },
        )
        for record in records
    ]


def group_emotions(
    records: list[EmotionRecord],
    emoji_by_group: dict[str, str] | None = None,
) -> list[EmotionGroup]:
    emoji_by_group = emoji_by_group or {}
    grouped: dict[str, list[EmotionOption]] = {}
    for record in records:
        grouped.setdefault(record.group, []).append(
            EmotionOption(
                emotion=record.emotion,
                reason=record.reason,
                advice=record.advice,
                services=record.services,
            )
        )
    return [
        EmotionGroup(
            group=group,
            emoji=emoji_by_group.get(group.casefold(), ""),
            options=options,
        )
        for group, options in grouped.items()
    ]
